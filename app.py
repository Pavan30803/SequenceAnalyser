from pathlib import Path

from flask import Flask, request, jsonify, send_file
import pandas as pd
import base64
import io
import json
import re
import zipfile
from datetime import datetime
from openpyxl import load_workbook
from openpyxl.styles.colors import COLOR_INDEX
from werkzeug.utils import secure_filename
from PIL import Image, ImageDraw, ImageFont

app = Flask(__name__)
app.config['MAX_CONTENT_LENGTH'] = 512 * 1024 * 1024
app.config['MAX_FORM_MEMORY_SIZE'] = 128 * 1024 * 1024
app.config['MAX_FORM_PARTS'] = 5000
BASE_DIR = Path(__file__).resolve().parent
FRONTEND_DIST_DIR = BASE_DIR / 'frontend' / 'dist'
BACKUP_DIR = BASE_DIR / 'Sequence_Backup'

# --- ANTI-CACHING BLOCK FOR DEVELOPMENT ---
@app.after_request
def add_header(response):
    response.headers['Cache-Control'] = 'no-store, no-cache, must-revalidate, post-check=0, pre-check=0, max-age=0'
    response.headers['Pragma'] = 'no-cache'
    response.headers['Expires'] = '-1'
    return response
# ------------------------------------------


@app.errorhandler(413)
def request_entity_too_large(error):
    return jsonify({'error': 'Backup request is too large. Reduce uploaded source files or save fewer attachments.'}), 413

def parse_excel(file_bytes):
    try:
        df = pd.read_excel(io.BytesIO(file_bytes), header=0)
    except Exception:
        last_error = None
        for encoding in ('utf-8-sig', 'utf-8', 'cp1252', 'latin1'):
            try:
                df = pd.read_csv(io.BytesIO(file_bytes), header=0, encoding=encoding)
                break
            except UnicodeDecodeError as error:
                last_error = error
        else:
            raise last_error
        
    df.columns = df.columns.astype(str).str.strip()
    return df


def sanitize_backup_name(value, fallback='file'):
    filename = secure_filename(str(value or '').strip())
    return filename or fallback


def parse_json_field(name, fallback):
    raw_value = request.form.get(name)
    if not raw_value:
        return fallback
    try:
        return json.loads(raw_value)
    except json.JSONDecodeError:
        return fallback


def read_json_file(path, fallback):
    if not path.exists():
        return fallback
    try:
        return json.loads(path.read_text(encoding='utf-8'))
    except (OSError, json.JSONDecodeError):
        return fallback


def get_backup_folders():
    if not BACKUP_DIR.exists():
        return []

    folders = []
    for folder in BACKUP_DIR.iterdir():
        if folder.is_dir() and ((folder / 'Mapped_Day_Opening_Report.csv').exists() or (folder / 'Mapped_Day_Opening_Report.xlsx').exists()):
            folders.append(folder)
    return sorted(folders, key=lambda item: item.stat().st_mtime, reverse=True)


def describe_backup_folder(folder):
    source_folder = folder / 'source_files'
    mapped_csv = folder / 'Mapped_Day_Opening_Report.csv'
    mapped_excel = folder / 'Mapped_Day_Opening_Report.xlsx'
    session_state = read_json_file(folder / 'Planner_Session.json', {})

    files = []
    for path in sorted(folder.rglob('*')):
        if path.is_file():
            files.append(str(path.relative_to(folder)).replace('\\', '/'))

    return {
        'id': folder.name,
        'folder': str(folder),
        'updated_at': datetime.fromtimestamp(folder.stat().st_mtime).isoformat(),
        'line_type': session_state.get('lineType', 'HDT'),
        'has_session_state': bool(session_state),
        'has_mapped_csv': mapped_csv.exists(),
        'has_mapped_excel': mapped_excel.exists(),
        'source_file_count': len([path for path in source_folder.iterdir() if path.is_file()]) if source_folder.exists() else 0,
        'files': files,
    }


def pdf_escape(value):
    return str(value).replace('\\', '\\\\').replace('(', '\\(').replace(')', '\\)')


def wrap_pdf_line(text, width=96):
    words = str(text).split()
    if not words:
        return ['']
    lines = []
    current = ''
    for word in words:
        candidate = f"{current} {word}".strip()
        if len(candidate) > width and current:
            lines.append(current)
            current = word
        else:
            current = candidate
    lines.append(current)
    return lines


def create_text_pdf(title, sections):
    lines = [title, f"Generated: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}", ""]
    for section_title, section_lines in sections:
        lines.extend([section_title, "-" * min(len(section_title), 80)])
        for line in section_lines:
            lines.extend(wrap_pdf_line(line))
        lines.append("")

    page_size = 48
    pages = [lines[index:index + page_size] for index in range(0, len(lines), page_size)] or [[]]
    objects = []
    pages_kids = []

    def add_object(content):
        objects.append(content)
        return len(objects)

    catalog_id = add_object("<< /Type /Catalog /Pages 2 0 R >>")
    pages_id = add_object("")
    font_id = add_object("<< /Type /Font /Subtype /Type1 /BaseFont /Helvetica >>")

    for page_lines in pages:
        text_commands = ["BT", "/F1 10 Tf", "50 790 Td", "14 TL"]
        for line in page_lines:
            text_commands.append(f"({pdf_escape(line)}) Tj")
            text_commands.append("T*")
        text_commands.append("ET")
        stream = "\n".join(text_commands)
        content_id = add_object(f"<< /Length {len(stream.encode('utf-8'))} >>\nstream\n{stream}\nendstream")
        page_id = add_object(
            f"<< /Type /Page /Parent {pages_id} 0 R /MediaBox [0 0 612 842] "
            f"/Resources << /Font << /F1 {font_id} 0 R >> >> /Contents {content_id} 0 R >>"
        )
        pages_kids.append(page_id)

    objects[pages_id - 1] = f"<< /Type /Pages /Kids [{' '.join(f'{kid} 0 R' for kid in pages_kids)}] /Count {len(pages_kids)} >>"

    pdf_parts = ["%PDF-1.4\n"]
    offsets = [0]
    for index, content in enumerate(objects, start=1):
        offsets.append(sum(len(part.encode('utf-8')) for part in pdf_parts))
        pdf_parts.append(f"{index} 0 obj\n{content}\nendobj\n")

    xref_offset = sum(len(part.encode('utf-8')) for part in pdf_parts)
    pdf_parts.append(f"xref\n0 {len(objects) + 1}\n")
    pdf_parts.append("0000000000 65535 f \n")
    for offset in offsets[1:]:
        pdf_parts.append(f"{offset:010d} 00000 n \n")
    pdf_parts.append(f"trailer\n<< /Size {len(objects) + 1} /Root {catalog_id} 0 R >>\nstartxref\n{xref_offset}\n%%EOF")
    return "".join(pdf_parts).encode('utf-8')


def flatten_count_dict(title, values):
    if not isinstance(values, dict) or not values:
        return [f"{title}: No data"]
    return [f"{title}: {key} = {value}" for key, value in sorted(values.items())]


def load_backup_font(size, bold=False):
    candidates = [
        "C:/Windows/Fonts/arialbd.ttf" if bold else "C:/Windows/Fonts/arial.ttf",
        "C:/Windows/Fonts/segoeuib.ttf" if bold else "C:/Windows/Fonts/segoeui.ttf",
    ]
    for candidate in candidates:
        try:
            return ImageFont.truetype(candidate, size=size)
        except Exception:
            continue
    return ImageFont.load_default()


def parse_chart_images(chart_images):
    parsed = []
    for item in chart_images if isinstance(chart_images, list) else []:
        if not isinstance(item, dict):
            continue
        image_data = str(item.get('image', ''))
        if ',' in image_data:
            header, image_data = image_data.split(',', 1)
        else:
            header = ''
        try:
            raw_bytes = base64.b64decode(image_data)
            image = Image.open(io.BytesIO(raw_bytes)).convert('RGB')
        except Exception:
            continue
        extension = 'jpg' if 'jpeg' in header.lower() or 'jpg' in header.lower() else 'png'
        parsed.append({
            'title': str(item.get('title') or 'Chart'),
            'image': image,
            'bytes': raw_bytes,
            'extension': extension,
        })
    return parsed


def draw_wrapped_text(draw, text, xy, font, fill=(25, 46, 74), max_width=1100, line_gap=8):
    x, y = xy
    words = str(text).split()
    if not words:
        return y + font.size + line_gap

    line = ''
    for word in words:
        candidate = f"{line} {word}".strip()
        bbox = draw.textbbox((x, y), candidate, font=font)
        if bbox[2] - bbox[0] > max_width and line:
            draw.text((x, y), line, font=font, fill=fill)
            y += font.size + line_gap
            line = word
        else:
            line = candidate
    draw.text((x, y), line, font=font, fill=fill)
    return y + font.size + line_gap


def make_backup_page(title=None):
    page = Image.new('RGB', (1240, 1754), 'white')
    draw = ImageDraw.Draw(page)
    if title:
        draw.text((70, 60), title, font=load_backup_font(34, True), fill=(0, 48, 98))
        draw.line((70, 112, 1170, 112), fill=(204, 218, 235), width=3)
    return page, draw


def paste_fit(page, image, box):
    x, y, width, height = box
    source = image.copy()
    source.thumbnail((width, height), Image.LANCZOS)
    paste_x = x + max((width - source.width) // 2, 0)
    paste_y = y + max((height - source.height) // 2, 0)
    page.paste(source, (paste_x, paste_y))


def build_backup_text_sections(line_type, summary, status_summary, inference_cards, mapped_row_count):
    sections = []
    sections.append((
        "Opening Summary",
        [
            f"Line type: {line_type}",
            f"Mapped report rows: {mapped_row_count}",
            f"Hold orders: {summary.get('total_hold', 0) if isinstance(summary, dict) else 0}",
            f"Skip vehicles: {summary.get('total_skipped', 0) if isinstance(summary, dict) else 0}",
        ],
    ))

    if isinstance(summary, dict):
        chart_lines = []
        chart_lines.extend(flatten_count_dict("Hold by model", summary.get('hold_stratification')))
        chart_lines.extend(flatten_count_dict("Skip by model", summary.get('skip_stratification')))
        chart_lines.extend(flatten_count_dict("Hold by type", summary.get('hold_type_stratification')))
        chart_lines.extend(flatten_count_dict("Skip by type", summary.get('skip_type_stratification')))
        chart_lines.extend(flatten_count_dict("Hold by work content", summary.get('hold_wc_stratification')))
        chart_lines.extend(flatten_count_dict("Skip by work content", summary.get('skip_wc_stratification')))
        chart_lines.extend(flatten_count_dict("Hold by region", summary.get('hold_region_stratification')))
        chart_lines.extend(flatten_count_dict("Skip by region", summary.get('skip_region_stratification')))
        sections.append(("Charts", chart_lines))

    status_lines = []
    if isinstance(status_summary, dict):
        status_lines.append(f"Production date: {status_summary.get('currentDay', '')}")
        status_lines.append(f"Vehicles in sequence: {status_summary.get('rowCount', 0)}")
        for shift in status_summary.get('shifts', []) or []:
            status_lines.append(f"{shift.get('label', 'Shift')}: {shift.get('rowCount', 0)} Vehicles in sequence")
            for status_name in ['engine', 'transmission', 'axle', 'frame']:
                status_lines.extend(flatten_count_dict(f"  {status_name.title()}", shift.get(status_name)))
        if not status_summary.get('shifts'):
            for status_name in ['engine', 'transmission', 'axle', 'frame']:
                status_lines.extend(flatten_count_dict(status_name.title(), status_summary.get(status_name)))
    sections.append(("Engine / Transmission / Axle / Frame Coverage", status_lines or ["No status summary available"]))

    shortage_lines = []
    for card in inference_cards if isinstance(inference_cards, list) else []:
        label = f"{card.get('part', '')} {card.get('partName', '')}".strip()
        if card.get('covered'):
            shortage_lines.append(f"{label}: Covered")
        else:
            shortage_lines.append(
                f"{label}: First shortage {card.get('shortageDate', '')}; "
                f"scheduled impact {card.get('scheduledImpactTime', card.get('impactTime', ''))}; "
                f"point-of-fit impact {card.get('pointOfFitImpactTime', 'N/A')} "
                f"at station {card.get('pointOfFitStation', 'N/A') or 'N/A'}; "
                f"sequence {card.get('firstDaySequences', '')}; models {card.get('connectingModels', '')}"
            )
            for entry in card.get('forecast', []) or []:
                shortage_lines.append(
                    f"  {entry.get('date', '')}: Day plan {entry.get('dayPlan', 0)}, Shortage qty {entry.get('shortageQty', 0)}"
                )
    sections.append(("Shortage Impact Analysis", shortage_lines or ["No shortage impact cards available"]))

    return sections


def group_chart_images(parsed_images):
    hold_images = [item for item in parsed_images if str(item.get('title', '')).lower().startswith('stratification: hold') or str(item.get('title', '')).lower().startswith('hold:')]
    skip_images = [item for item in parsed_images if str(item.get('title', '')).lower().startswith('stratification: skip') or str(item.get('title', '')).lower().startswith('skip:')]
    other_images = [item for item in parsed_images if item not in hold_images and item not in skip_images]
    return [
        ("Hold PI Chart & Further Stratifications", hold_images),
        ("Skip PI Chart & Further Stratifications", skip_images),
        ("Additional Stratifications", other_images),
    ]


def add_chart_group_page(pages, title, chart_group):
    if not chart_group:
        return
    page, draw = make_backup_page(title)
    title_font = load_backup_font(22, True)
    for index, item in enumerate(chart_group[:4]):
        if index == 0:
            box = (80, 165, 1080, 760)
            label_xy = (90, 135)
        else:
            column_width = 340
            x = 80 + (index - 1) * 370
            box = (x, 1045, column_width, 430)
            label_xy = (x, 1010)
        draw.text(label_xy, item['title'], font=title_font, fill=(0, 48, 98))
        paste_fit(page, item['image'], box)
    pages.append(page)


def build_backup_pdf_bytes(line_type, summary, status_summary, inference_cards, mapped_row_count, chart_images=None):
    sections = build_backup_text_sections(line_type, summary, status_summary, inference_cards, mapped_row_count)
    parsed_images = parse_chart_images(chart_images)
    if not parsed_images:
        return create_text_pdf("Day Opening Constraint Backup", sections)

    pages = []
    page, draw = make_backup_page("Day Opening Constraint Backup")
    normal_font = load_backup_font(20)
    section_font = load_backup_font(24, True)
    y = 135
    for section_title, lines in sections:
        if y > 1500:
            pages.append(page)
            page, draw = make_backup_page()
            y = 70
        draw.text((70, y), section_title, font=section_font, fill=(0, 48, 98))
        y += 42
        for line in lines[:18]:
            y = draw_wrapped_text(draw, line, (90, y), normal_font, max_width=1060)
            if y > 1540:
                pages.append(page)
                page, draw = make_backup_page()
                y = 70
        y += 20
    pages.append(page)

    for title, group in group_chart_images(parsed_images):
        add_chart_group_page(pages, title, group)

    output = io.BytesIO()
    pages[0].save(output, format='PDF', save_all=True, append_images=pages[1:])
    return output.getvalue()


def xml_escape(value):
    return str(value).replace('&', '&amp;').replace('<', '&lt;').replace('>', '&gt;').replace('"', '&quot;')


def build_backup_pptx_bytes(line_type, summary, status_summary, inference_cards, mapped_row_count, chart_images=None):
    parsed_images = parse_chart_images(chart_images)
    sections = build_backup_text_sections(line_type, summary, status_summary, inference_cards, mapped_row_count)
    slide_titles = ["Day Opening Constraint Backup"]
    slide_payloads = [{"type": "text", "sections": sections[:2]}]
    for title, group in group_chart_images(parsed_images):
        if group:
            slide_titles.append(title)
            slide_payloads.append({"type": "charts", "charts": group[:4]})
    slide_titles.append("Shortage Impact Analysis")
    slide_payloads.append({"type": "text", "sections": [sections[-1]]})

    media_items = []
    slide_xml = []
    slide_rels = []
    for slide_index, (title, payload) in enumerate(zip(slide_titles, slide_payloads), start=1):
        shapes = [
            f'<p:sp><p:nvSpPr><p:cNvPr id="2" name="Title"/><p:cNvSpPr/><p:nvPr/></p:nvSpPr><p:spPr><a:xfrm><a:off x="420000" y="260000"/><a:ext cx="11300000" cy="520000"/></a:xfrm></p:spPr><p:txBody><a:bodyPr/><a:lstStyle/><a:p><a:r><a:rPr lang="en-US" sz="3000" b="1"/><a:t>{xml_escape(title)}</a:t></a:r></a:p></p:txBody></p:sp>'
        ]
        rels = []
        if payload["type"] == "charts":
            boxes = [
                (450000, 930000, 5300000, 3900000),
                (6050000, 930000, 2600000, 1950000),
                (8850000, 930000, 2600000, 1950000),
                (6050000, 3230000, 5400000, 2400000),
            ]
            for chart_index, chart in enumerate(payload["charts"], start=1):
                media_name = f"image{len(media_items) + 1}.{chart.get('extension', 'png')}"
                media_items.append((media_name, chart["bytes"]))
                rel_id = f"rId{chart_index}"
                rels.append((rel_id, f"../media/{media_name}"))
                x, y, cx, cy = boxes[chart_index - 1]
                shapes.append(
                    f'<p:pic><p:nvPicPr><p:cNvPr id="{chart_index + 2}" name="{xml_escape(chart["title"])}"/><p:cNvPicPr/><p:nvPr/></p:nvPicPr><p:blipFill><a:blip r:embed="{rel_id}"/><a:stretch><a:fillRect/></a:stretch></p:blipFill><p:spPr><a:xfrm><a:off x="{x}" y="{y}"/><a:ext cx="{cx}" cy="{cy}"/></a:xfrm><a:prstGeom prst="rect"><a:avLst/></a:prstGeom></p:spPr></p:pic>'
                )
        else:
            y = 980000
            body_lines = []
            for section_title, lines in payload["sections"]:
                body_lines.append(section_title)
                body_lines.extend(lines[:10])
            for line in body_lines[:22]:
                shapes.append(
                    f'<p:sp><p:nvSpPr><p:cNvPr id="{len(shapes) + 2}" name="Text"/><p:cNvSpPr/><p:nvPr/></p:nvSpPr><p:spPr><a:xfrm><a:off x="620000" y="{y}"/><a:ext cx="10800000" cy="270000"/></a:xfrm></p:spPr><p:txBody><a:bodyPr/><a:lstStyle/><a:p><a:r><a:rPr lang="en-US" sz="1400"/><a:t>{xml_escape(line)}</a:t></a:r></a:p></p:txBody></p:sp>'
                )
                y += 260000

        slide_xml.append(f'<?xml version="1.0" encoding="UTF-8" standalone="yes"?><p:sld xmlns:a="http://schemas.openxmlformats.org/drawingml/2006/main" xmlns:r="http://schemas.openxmlformats.org/officeDocument/2006/relationships" xmlns:p="http://schemas.openxmlformats.org/presentationml/2006/main"><p:cSld><p:spTree><p:nvGrpSpPr><p:cNvPr id="1" name=""/><p:cNvGrpSpPr/><p:nvPr/></p:nvGrpSpPr><p:grpSpPr><a:xfrm><a:off x="0" y="0"/><a:ext cx="0" cy="0"/><a:chOff x="0" y="0"/><a:chExt cx="0" cy="0"/></a:xfrm></p:grpSpPr>{"".join(shapes)}</p:spTree></p:cSld></p:sld>')
        slide_rels.append(rels)

    output = io.BytesIO()
    with zipfile.ZipFile(output, 'w', zipfile.ZIP_DEFLATED) as archive:
        overrides = ''.join(f'<Override PartName="/ppt/slides/slide{i}.xml" ContentType="application/vnd.openxmlformats-officedocument.presentationml.slide+xml"/>' for i in range(1, len(slide_xml) + 1))
        archive.writestr('[Content_Types].xml', f'<?xml version="1.0" encoding="UTF-8" standalone="yes"?><Types xmlns="http://schemas.openxmlformats.org/package/2006/content-types"><Default Extension="rels" ContentType="application/vnd.openxmlformats-package.relationships+xml"/><Default Extension="xml" ContentType="application/xml"/><Default Extension="png" ContentType="image/png"/><Default Extension="jpg" ContentType="image/jpeg"/><Default Extension="jpeg" ContentType="image/jpeg"/><Override PartName="/ppt/presentation.xml" ContentType="application/vnd.openxmlformats-officedocument.presentationml.presentation.main+xml"/>{overrides}</Types>')
        archive.writestr('_rels/.rels', '<?xml version="1.0" encoding="UTF-8" standalone="yes"?><Relationships xmlns="http://schemas.openxmlformats.org/package/2006/relationships"><Relationship Id="rId1" Type="http://schemas.openxmlformats.org/officeDocument/2006/relationships/officeDocument" Target="ppt/presentation.xml"/></Relationships>')
        presentation_rels = ''.join(f'<Relationship Id="rId{i}" Type="http://schemas.openxmlformats.org/officeDocument/2006/relationships/slide" Target="slides/slide{i}.xml"/>' for i in range(1, len(slide_xml) + 1))
        slide_ids = ''.join(f'<p:sldId id="{255 + i}" r:id="rId{i}"/>' for i in range(1, len(slide_xml) + 1))
        archive.writestr('ppt/presentation.xml', f'<?xml version="1.0" encoding="UTF-8" standalone="yes"?><p:presentation xmlns:a="http://schemas.openxmlformats.org/drawingml/2006/main" xmlns:r="http://schemas.openxmlformats.org/officeDocument/2006/relationships" xmlns:p="http://schemas.openxmlformats.org/presentationml/2006/main"><p:sldSz cx="12192000" cy="6858000" type="wide"/><p:sldIdLst>{slide_ids}</p:sldIdLst></p:presentation>')
        archive.writestr('ppt/_rels/presentation.xml.rels', f'<?xml version="1.0" encoding="UTF-8" standalone="yes"?><Relationships xmlns="http://schemas.openxmlformats.org/package/2006/relationships">{presentation_rels}</Relationships>')
        for index, xml in enumerate(slide_xml, start=1):
            archive.writestr(f'ppt/slides/slide{index}.xml', xml)
            rel_xml = ''.join(f'<Relationship Id="{rel_id}" Type="http://schemas.openxmlformats.org/officeDocument/2006/relationships/image" Target="{target}"/>' for rel_id, target in slide_rels[index - 1])
            archive.writestr(f'ppt/slides/_rels/slide{index}.xml.rels', f'<?xml version="1.0" encoding="UTF-8" standalone="yes"?><Relationships xmlns="http://schemas.openxmlformats.org/package/2006/relationships">{rel_xml}</Relationships>')
        for media_name, image_bytes in media_items:
            archive.writestr(f'ppt/media/{media_name}', image_bytes)
    return output.getvalue()


def save_uploaded_file(file_storage, target_dir, prefix):
    if not file_storage or not file_storage.filename:
        return None
    safe_name = sanitize_backup_name(file_storage.filename, f"{prefix}.dat")
    target = target_dir / f"{prefix}_{safe_name}"
    file_storage.save(target)
    return target.name

def normalize_order_key(value):
    text = str(value).strip()
    if not text or text.lower() == 'nan':
        return ''
    text = text.replace('\u00a0', '').replace(' ', '')
    if text.endswith('.0'):
        text = text[:-2]
    digits = ''.join(ch for ch in text if ch.isdigit())
    return digits[-6:] if len(digits) >= 6 else digits


def normalize_fill_hex(cell):
    for color in (cell.fill.fgColor, cell.fill.start_color, cell.fill.bgColor):
        if color.type == 'rgb' and color.rgb:
            return str(color.rgb)[-6:].upper()
        if color.type == 'indexed' and color.indexed is not None:
            indexed_color = COLOR_INDEX[color.indexed]
            if indexed_color:
                return str(indexed_color)[-6:].upper()
    return ''


def combine_axle_status(colors):
    color_set = {color for color in colors if color}
    if not color_set:
        return ''
    if 'FF9900' in color_set:
        return 'WIP'
    if 'C0C0C0' in color_set:
        return 'NOT STARTED'
    if color_set == {'00FF00'}:
        return 'AVAILABLE'
    if color_set.issubset({'00FF00', 'FFFF00'}):
        return 'IN TRANSIT'
    return ''


def parse_axle_status_report(file_bytes):
    workbook = load_workbook(io.BytesIO(file_bytes), data_only=True)
    sheet = workbook.active
    order_colors = {}

    for row in sheet.iter_rows(min_row=2):
        if len(row) < 5:
            continue
        order_key = normalize_order_key(row[1].value)
        if not order_key:
            continue
        color_hex = normalize_fill_hex(row[4])
        if color_hex:
            order_colors.setdefault(order_key, []).append(color_hex)

    return {
        order_key: combine_axle_status(colors)
        for order_key, colors in order_colors.items()
        if combine_axle_status(colors)
    }


def parse_frame_status_report(file_bytes):
    workbook = load_workbook(io.BytesIO(file_bytes), data_only=True)
    sheet = workbook['Sheet1'] if 'Sheet1' in workbook.sheetnames else workbook['Sheet 1'] if 'Sheet 1' in workbook.sheetnames else workbook.active
    frame_statuses = {}

    for row in sheet.iter_rows(min_row=2):
        if len(row) < 30:
            continue
        dsn_key = normalize_order_key(row[2].value)
        if not dsn_key:
            continue
        status = str(row[29].value or '').strip()
        status_key = status.upper()
        part_shortage_note = str(row[30].value or '').strip() if len(row) > 30 else ''
        part_shortage_key = part_shortage_note.upper()
        if status_key in {'DICV DOL', 'TRANSIT', 'SMS FG'}:
            status = 'COVERED'
        elif status_key == 'TO BE PROD' and 'PART SHORTAGE' in part_shortage_key:
            status = 'Part Shortage'
        if status:
            frame_statuses[dsn_key] = status

    return frame_statuses


def normalize_part_number_value(value):
    if pd.isna(value):
        return ''
    if isinstance(value, float) and value.is_integer():
        return str(int(value)).strip().replace('.', '')
    return str(value).strip().replace('.', '')


def normalize_text_value(value):
    if pd.isna(value):
        return ''
    if isinstance(value, float) and value.is_integer():
        return str(int(value)).strip()
    return str(value).strip()


def normalize_detail_value(value):
    normalized = normalize_text_value(value)
    return normalized if normalized else 'NA'


def normalize_requirement_qty(value):
    if pd.isna(value):
        return ''
    if isinstance(value, float) and value.is_integer():
        return int(value)
    return str(value).strip()


def get_requirement_status(cell):
    color = normalize_fill_hex(cell)
    status_by_color = {
        'D0021B': 'SHORTAGE',
        '7ED321': 'AVAILABLE',
        '9B9B9B': 'NOT IN DEMAND',
    }
    return status_by_color.get(color, '')


def aggregate_requirement_status(statuses):
    normalized_statuses = {status for status in statuses if status}
    if 'SHORTAGE' in normalized_statuses:
        return 'SHORTAGE'
    if 'AVAILABLE' in normalized_statuses:
        return 'AVAILABLE'
    if 'NOT IN DEMAND' in normalized_statuses:
        return 'NOT IN DEMAND'
    return ''


def build_requirement_coverage_from_sheet(row):
    coverage_config = [
        ('N', 29, 30),
        ('N+1', 33, 34),
        ('N+2', 37, 38),
        ('N+3', 41, 42),
        ('N+4', 45, 46),
    ]
    coverage = []
    for day_label, a_index, b_index in coverage_config:
        a_cell = row[a_index] if len(row) > a_index else None
        b_cell = row[b_index] if len(row) > b_index else None
        a_status = get_requirement_status(a_cell) if a_cell else ''
        b_status = get_requirement_status(b_cell) if b_cell else ''
        coverage.append({
            'day': day_label,
            'dayStatus': aggregate_requirement_status([a_status, b_status]),
            'shifts': [
                {
                    'label': 'A Shift',
                    'qty': normalize_requirement_qty(a_cell.value) if a_cell else '',
                    'status': a_status,
                },
                {
                    'label': 'B Shift',
                    'qty': normalize_requirement_qty(b_cell.value) if b_cell else '',
                    'status': b_status,
                },
            ],
        })
    return coverage


def build_requirement_coverage_from_values(row_values):
    coverage_config = [
        ('N', 29, 30),
        ('N+1', 33, 34),
        ('N+2', 37, 38),
        ('N+3', 41, 42),
        ('N+4', 45, 46),
    ]
    coverage = []
    for day_label, a_index, b_index in coverage_config:
        a_value = row_values[a_index] if len(row_values) > a_index else ''
        b_value = row_values[b_index] if len(row_values) > b_index else ''
        coverage.append({
            'day': day_label,
            'dayStatus': '',
            'shifts': [
                {
                    'label': 'A Shift',
                    'qty': normalize_requirement_qty(a_value),
                    'status': '',
                },
                {
                    'label': 'B Shift',
                    'qty': normalize_requirement_qty(b_value),
                    'status': '',
                },
            ],
        })
    return coverage


def build_critical_part_detail_from_values(row_values, current_part):
    return {
        'partNumber': current_part,
        'partDescription': normalize_detail_value(row_values[1]),
        'vendorName': normalize_detail_value(row_values[4]),
        'supplierBacklog': normalize_detail_value(row_values[23] if len(row_values) > 23 else ''),
        'l4Name': normalize_detail_value(row_values[6]),
        'pmcName': normalize_detail_value(row_values[7]),
        'smName': normalize_detail_value(row_values[8] if len(row_values) > 8 else ''),
        'requirementCoverage': build_requirement_coverage_from_values(row_values),
    }


def build_critical_part_detail_from_sheet(row, current_part):
    return {
        'partNumber': current_part,
        'partDescription': normalize_detail_value(row[1].value),
        'vendorName': normalize_detail_value(row[4].value),
        'supplierBacklog': normalize_detail_value(row[23].value if len(row) > 23 else ''),
        'l4Name': normalize_detail_value(row[6].value),
        'pmcName': normalize_detail_value(row[7].value),
        'smName': normalize_detail_value(row[8].value if len(row) > 8 else ''),
        'requirementCoverage': build_requirement_coverage_from_sheet(row),
    }


def parse_critical_parts_details_from_table(file_bytes, part_numbers):
    df = parse_excel(file_bytes)
    target_parts = {normalize_part_number_value(part).upper() for part in part_numbers if normalize_part_number_value(part)}
    details_by_part = {}
    for _, row in df.iterrows():
        row_values = list(row)
        if len(row_values) < 8:
            continue
        current_part = normalize_part_number_value(row_values[0])
        current_key = current_part.upper()
        if current_key not in target_parts or current_key in details_by_part:
            continue
        details_by_part[current_key] = build_critical_part_detail_from_values(row_values, current_part)
        if len(details_by_part) == len(target_parts):
            break
    return details_by_part


def parse_critical_parts_details(file_bytes, part_numbers):
    target_parts = {normalize_part_number_value(part).upper() for part in part_numbers if normalize_part_number_value(part)}
    if not target_parts:
        return {}

    try:
        workbook = load_workbook(io.BytesIO(file_bytes), data_only=True)
    except Exception:
        return parse_critical_parts_details_from_table(file_bytes, part_numbers)

    sheet = workbook.active
    details_by_part = {}
    for row in sheet.iter_rows(min_row=2):
        if len(row) < 8:
            continue
        current_part = normalize_part_number_value(row[0].value)
        current_key = current_part.upper()
        if current_key not in target_parts or current_key in details_by_part:
            continue
        details_by_part[current_key] = build_critical_part_detail_from_sheet(row, current_part)
        if len(details_by_part) == len(target_parts):
            break

    return details_by_part


def parse_critical_part_details(file_bytes, part_number):
    return parse_critical_parts_details(file_bytes, [part_number]).get(normalize_part_number_value(part_number).upper())


def build_l4_directory_entry(l4_value, pmc_value, sm_value):
    l4_name = normalize_detail_value(l4_value)
    if l4_name == 'NA':
        return None
    return {
        'l4Name': l4_name,
        'pmcName': normalize_detail_value(pmc_value),
        'smName': normalize_detail_value(sm_value),
    }


def parse_l4_directory_from_table(file_bytes):
    df = parse_excel(file_bytes)
    directory = {}
    for _, row in df.iterrows():
        row_values = list(row)
        if len(row_values) < 9:
            continue
        entry = build_l4_directory_entry(row_values[6], row_values[7], row_values[8])
        if not entry:
            continue
        group = directory.setdefault(entry['l4Name'], {'l4Name': entry['l4Name'], 'pmcNames': set(), 'smNames': set()})
        if entry['pmcName'] != 'NA':
            group['pmcNames'].add(entry['pmcName'])
        if entry['smName'] != 'NA':
            group['smNames'].add(entry['smName'])
    return [
        {
            'l4Name': group['l4Name'],
            'pmcNames': sorted(group['pmcNames']),
            'smNames': sorted(group['smNames']),
        }
        for group in sorted(directory.values(), key=lambda item: item['l4Name'])
    ]


def parse_l4_directory(file_bytes):
    try:
        workbook = load_workbook(io.BytesIO(file_bytes), data_only=True)
    except Exception:
        return parse_l4_directory_from_table(file_bytes)

    sheet = workbook.active
    directory = {}
    for row in sheet.iter_rows(min_row=2):
        if len(row) < 9:
            continue
        entry = build_l4_directory_entry(row[6].value, row[7].value, row[8].value)
        if not entry:
            continue
        group = directory.setdefault(entry['l4Name'], {'l4Name': entry['l4Name'], 'pmcNames': set(), 'smNames': set()})
        if entry['pmcName'] != 'NA':
            group['pmcNames'].add(entry['pmcName'])
        if entry['smName'] != 'NA':
            group['smNames'].add(entry['smName'])
    return [
        {
            'l4Name': group['l4Name'],
            'pmcNames': sorted(group['pmcNames']),
            'smNames': sorted(group['smNames']),
        }
        for group in sorted(directory.values(), key=lambda item: item['l4Name'])
    ]


def analyze_sequence(df, shortages, engine_transmission_statuses=None, axle_statuses=None, frame_statuses=None, opening_hold_keys=None, line_type='HDT'):
    opening_hold_keys = {normalize_order_key(key) for key in (opening_hold_keys or []) if normalize_order_key(key)}
    line_type = str(line_type or 'HDT').strip().upper()

    if df.empty:
        return {
            'summary': { 
                'dsn_min': 0, 'dsn_max': 0, 'total_in_file': 0, 'total_hold': 0, 'total_skipped': 0, 
                'state_distribution': {}, 'hold_stratification': {}, 'skip_stratification': {}, 
                'hold_type_stratification': {}, 'skip_type_stratification': {}, 
                'hold_wc_stratification': {}, 'skip_wc_stratification': {},
                'hold_region_stratification': {}, 'skip_region_stratification': {},
                'shortage_parts': [] 
            },
            'hold_orders': [], 'skip_orders': [], 'gaps': [], 'preview_columns': [], 'preview_data': []
        }

    original_df = df.copy()

    # DYNAMIC COLUMN IDENTIFICATION
    dsn_col = next((c for c in original_df.columns if str(c).strip().upper() in ['DSN', 'DELIVERY SEQUENCE NUMBER']), None)
    if not dsn_col and len(original_df.columns) > 2:
        dsn_col = original_df.columns[2]

    serial_col = next((c for c in original_df.columns if str(c).strip().upper() == 'SERIAL NUMBER'), None)
    if not serial_col and len(original_df.columns) > 3:
        serial_col = original_df.columns[3]
        
    variant_col = next((c for c in original_df.columns if str(c).strip().upper() == 'VARIANT'), None)
    if not variant_col and len(original_df.columns) > 6:
        variant_col = original_df.columns[6]

    desc_col = next((c for c in original_df.columns if str(c).strip().upper() == 'DESCRIPTION'), None)
    if not desc_col and len(original_df.columns) > 7:
        desc_col = original_df.columns[7]
        
    status_col = next((c for c in original_df.columns if str(c).strip().upper() == 'STATUS'), None)
    if not status_col and len(original_df.columns) > 8:
        status_col = original_df.columns[8]
        
    state_col = next((c for c in original_df.columns if str(c).strip().upper() in ['VEHICLE ORDER STATE', 'STATE']), None)
    if not state_col and len(original_df.columns) > 9:
        state_col = original_df.columns[9]
        
    order_col = next((c for c in original_df.columns if str(c).strip().upper() in ['ORDER NUMBER', 'ORDER NO']), None)

    if not serial_col or not state_col:
        raise ValueError("Required columns (Column D, Column I, or Column J) could not be found in the uploaded file.")

    def get_vehicle_key(row):
        if order_col:
            order_key = normalize_order_key(row.get(order_col, ''))
            if order_key:
                return order_key
        serial_key = normalize_order_key(row.get(serial_col, '')) if serial_col else ''
        if serial_key:
            return serial_key
        return normalize_order_key(row.get(dsn_col, '')) if dsn_col else ''

    def get_vehicle_lookup_keys(row):
        keys = []
        candidate_cols = [order_col, serial_col, dsn_col]
        for fallback_idx in (1, 2, 3):
            if len(original_df.columns) > fallback_idx:
                candidate_cols.append(original_df.columns[fallback_idx])

        for column in candidate_cols:
            if not column:
                continue
            key = normalize_order_key(row.get(column, ''))
            if key and key not in keys:
                keys.append(key)

        return keys

    def get_first_mapped_value(mapping, row, default=''):
        for key in get_vehicle_lookup_keys(row):
            if key in mapping:
                return mapping[key]
        return default

    def is_released_hold_row(row):
        if not opening_hold_keys:
            return False
        vehicle_key = get_vehicle_key(row)
        current_state = str(row.get(state_col, '')).strip().upper()
        return bool(vehicle_key and vehicle_key in opening_hold_keys and current_state != 'HOLD')

    extra_cols = ['Order Number', 'Hold Status', 'Vehicle Start Time', 'Country']
    extra_cols = [c for c in extra_cols if c in original_df.columns]

    # =========================================================
    # LOGIC HELPERS: MODEL, WORK CONTENT, & REGION
    # =========================================================
    def is_bus_variant(variant_val):
        return str(variant_val).strip().upper().startswith(('V83', 'F83', 'M83', 'L83'))

    def extract_model(desc_val, variant_val=''):
        desc_str = str(desc_val).strip()
        if not desc_str or desc_str.lower() == 'nan':
            return 'Unknown'
        if line_type == 'MDT':
            if is_bus_variant(variant_val):
                return desc_str[:4] if len(desc_str) >= 4 else desc_str
            desc_upper = desc_str.upper()
            first_token = desc_str.split()[0] if desc_str.split() else ''
            if first_token.isdigit():
                return first_token
            if len(desc_upper) >= 6 and desc_upper[4:6] in ['RE', 'RD']:
                return desc_str[:6]
            if len(desc_upper) >= 5 and desc_upper[4] in ['R', 'C']:
                return desc_str[:5]
            return desc_str[:4] if len(desc_str) >= 4 else desc_str
        if len(desc_str) >= 6 and desc_str[5] in ['T', 'S', 'M']:
            return desc_str[:6]
        elif len(desc_str) >= 5:
            return desc_str[:5]
        return desc_str

    def get_work_content(variant_val, desc_val):
        if is_bus_variant(variant_val):
            return 'HWC'
        desc_str = str(desc_val).strip().upper()
        if len(desc_str) >= 5 and desc_str[4] == 'T' and '4X2' in desc_str:
            return 'LWC'
        if len(desc_str) >= 6 and desc_str[4:6] == 'CM':
            return 'HWC'
        if len(desc_str) >= 2:
            first_two = desc_str[:2]
            if first_two.isdigit() and int(first_two) > 30:
                return 'HWC'
        return 'LWC'
        
    def get_region(variant_val):
        if str(variant_val).strip().upper().startswith('V'):
            return 'Domestic'
        return 'Export'

    def get_pie_model_label(record):
        model = record.get('model', 'Unknown')
        if is_bus_variant(record.get('variant', '')):
            region = str(record.get('region', 'Export')).strip()
            return f"{model} (BUS)" if region == 'Domestic' else f"{model} (Exp BUS)"
        return model

    def get_engine_status_fallback(variant_val):
        variant_str = str(variant_val).strip().upper()
        if len(variant_str) >= 11 and variant_str[10] in ['U', 'T']:
            return 'Rapid Prime'
        return ''

    include_work_content = line_type != 'MDT'

    # Insert Data Columns into original_df
    original_df['Model'] = original_df.apply(lambda row: extract_model(row.get(desc_col, ''), row.get(variant_col, '')), axis=1)
    if include_work_content:
        original_df['Work Content'] = original_df.apply(lambda row: get_work_content(row.get(variant_col, ''), row.get(desc_col, '')), axis=1)
    original_df['Region'] = original_df[variant_col].apply(get_region)
    
    cols = list(original_df.columns)
    cols.remove('Model')
    if include_work_content:
        cols.remove('Work Content')
    cols.remove('Region')
        
    if desc_col in cols:
        desc_idx = cols.index(desc_col)
        cols.insert(desc_idx + 1, 'Model')
        if include_work_content:
            cols.insert(desc_idx + 2, 'Work Content')
            cols.insert(desc_idx + 3, 'Region')
        else:
            cols.insert(desc_idx + 2, 'Region')
    else:
        cols.append('Model')
        if include_work_content:
            cols.append('Work Content')
        cols.append('Region')
        
    original_df = original_df[cols]

    # =========================================================
    # ENGINE & TRANSMISSION STATUS MAPPING
    # =========================================================
    has_engine_transmission_report = engine_transmission_statuses is not None
    engine_transmission_statuses = engine_transmission_statuses or {}
    status_insert_idx = len(original_df.columns)
    hold_status_col = next((c for c in original_df.columns if str(c).strip().upper() == 'HOLD STATUS'), None)
    if hold_status_col:
        status_insert_idx = list(original_df.columns).index(hold_status_col) + 1

    if has_engine_transmission_report:
        engine_values = []
        transmission_values = []
        for _, row in original_df.iterrows():
            mapped_status = get_first_mapped_value(engine_transmission_statuses, row, {})
            engine_status = mapped_status.get('engine_status', '')
            if not engine_status:
                engine_status = get_engine_status_fallback(row.get(variant_col, '')) if variant_col else ''
            engine_values.append(engine_status)
            transmission_values.append(mapped_status.get('transmission_status', ''))

        original_df.insert(status_insert_idx, 'Engine status', engine_values)
        original_df.insert(status_insert_idx + 1, 'Transmission status', transmission_values)
        extra_cols.extend(['Engine status', 'Transmission status'])

    # =========================================================
    # AXLE STATUS MAPPING
    # =========================================================
    if axle_statuses:
        axle_values = []
        for _, row in original_df.iterrows():
            axle_values.append(get_first_mapped_value(axle_statuses, row))

        axle_insert_idx = len(original_df.columns)
        country_col = next((c for c in original_df.columns if str(c).strip().upper() == 'COUNTRY'), None)
        if country_col:
            axle_insert_idx = list(original_df.columns).index(country_col)

        original_df.insert(axle_insert_idx, 'Axle status', axle_values)
        extra_cols.append('Axle status')

    # =========================================================
    # FRAME STATUS MAPPING
    # =========================================================
    if frame_statuses:
        frame_values = []
        for _, row in original_df.iterrows():
            dsn_key = normalize_order_key(row.get(dsn_col, '')) if dsn_col else ''
            frame_values.append(frame_statuses.get(dsn_key, ''))

        frame_insert_idx = len(original_df.columns)
        country_col = next((c for c in original_df.columns if str(c).strip().upper() == 'COUNTRY'), None)
        if country_col:
            frame_insert_idx = list(original_df.columns).index(country_col)
        elif 'Region' in original_df.columns:
            frame_insert_idx = list(original_df.columns).index('Region')
        elif 'Axle status' in original_df.columns:
            frame_insert_idx = list(original_df.columns).index('Axle status') + 1

        original_df.insert(frame_insert_idx, 'Frame status', frame_values)
        extra_cols.append('Frame status')

    # =========================================================
    # PART SHORTAGE MAPPING LOGIC (WITH REF & QTY)
    # =========================================================
    cols = list(original_df.columns)
    insert_idx = len(cols)
    hold_status_col = next((c for c in cols if str(c).strip().upper() == 'HOLD STATUS'), None)
    if 'Transmission status' in cols:
        insert_idx = cols.index('Transmission status') + 1
    elif 'Engine status' in cols:
        insert_idx = cols.index('Engine status') + 1
    elif hold_status_col:
        insert_idx = cols.index(hold_status_col) + 1

    for part_num, details in shortages.items():
        var_set = details['variants']
        ref_order = details['ref_order']
        qty = details['qty']
        usage = max(int(details.get('usage', 1) or 1), 1)
        
        start_idx = 0
        if order_col and ref_order:
            matches = original_df.index[original_df[order_col].astype(str).str.strip() == ref_order].tolist()
            if matches:
                start_idx = matches[0]

        col_data = [''] * len(original_df)
        for i in range(len(original_df)):
            var = str(original_df.iloc[i].get(variant_col, '')).strip().upper()
            if var in var_set:
                if i < start_idx:
                    col_data[i] = 'Covered'
                else:
                    if qty >= usage:
                        col_data[i] = 'Covered'
                        qty -= usage
                    else:
                        col_data[i] = '⚠️ SHORTAGE'
        
        original_df.insert(insert_idx, part_num, col_data)
        insert_idx += 1
        extra_cols.append(part_num)

    def build_record_dict(row, seq_val):
        d = {
            'seq_val': seq_val,
            'dsn': str(row.get(dsn_col, '')),
            'serial': str(row.get(serial_col, '')),
            'vehicle_order_state': str(row.get(state_col, '')),
            'status': str(row.get(status_col, '')) if status_col else '',
            'description': str(row.get(desc_col, '')) if desc_col else '',
            'model': str(row.get('Model', '')),
            'variant': str(row.get(variant_col, '')) if variant_col else '',
            'work_content': str(row.get('Work Content', '')) if include_work_content else '',
            'region': str(row.get('Region', ''))
        }
        for c in extra_cols:
            val = row.get(c, '')
            d[c.lower().replace(' ', '_')] = '' if pd.isna(val) else str(val)
        return d

    # =========================================================
    # 1. HOLD LOGIC
    # =========================================================
    hold_mask = original_df[state_col].astype(str).str.strip().str.upper() == 'HOLD'
    hold_rows = original_df[hold_mask].copy()

    hold_records = []
    for _, r in hold_rows.iterrows():
        # Fallback seq_val for hold records
        raw_s = str(r.get(serial_col, '')).replace('.0', '').strip()
        seq_val = int(raw_s[-5:]) if len(raw_s) >= 5 and raw_s[-5:].isdigit() else 0
        hold_records.append(build_record_dict(r, seq_val))

    # =========================================================
    # 2. STRICT SEQUENCE SKIP LOGIC (Anomaly Block Tracking)
    # =========================================================
    df = original_df.copy()
    raw_serial = df[serial_col].astype(str).str.replace(r'\.0$', '', regex=True).str.strip()
    df['_seq_int'] = pd.to_numeric(raw_serial.str[-5:], errors='coerce')
    df = df.dropna(subset=['_seq_int'])
    df['_seq_int'] = df['_seq_int'].astype(int)

    gaps = []
    skip_records = []
    
    last_valid_seq = None
    is_in_anomaly_block = False
    current_anomaly_group = []
    anomaly_start_seq = None
    released_hold_anchor_seq = None

    for _, row in df.iterrows():
        current_seq = int(row['_seq_int'])
        status_val = str(row.get(status_col, '')).strip().upper() if status_col else ''

        if last_valid_seq is None:
            last_valid_seq = current_seq
            continue

        expected_next = last_valid_seq + 1
        released_hold_row = is_released_hold_row(row)

        if released_hold_anchor_seq is not None and current_seq == released_hold_anchor_seq + 1:
            last_valid_seq = current_seq
            released_hold_anchor_seq = None
            continue

        if current_seq == expected_next:
            last_valid_seq = current_seq
            released_hold_anchor_seq = None
            
            if is_in_anomaly_block:
                first_anomaly = current_anomaly_group[0]['_seq_int']
                last_anomaly = current_anomaly_group[-1]['_seq_int']
                skipped_range = f"{first_anomaly} – {last_anomaly}" if first_anomaly != last_anomaly else str(first_anomaly)

                gaps.append({
                    'from_dsn': anomaly_start_seq,
                    'to_dsn': current_seq,
                    'skipped_count': len(current_anomaly_group),
                    'skipped_range': skipped_range
                })
                
                current_anomaly_group = []
                is_in_anomaly_block = False
        else:
            if released_hold_row:
                released_hold_anchor_seq = current_seq
                continue

            if is_in_anomaly_block:
                current_anomaly_group.append(row)
                if status_val == 'TRIM LINE':
                    skip_records.append(build_record_dict(row, current_seq))
            else:
                if status_val == 'TRIM LINE':
                    is_in_anomaly_block = True
                    anomaly_start_seq = last_valid_seq
                    current_anomaly_group.append(row)
                    skip_records.append(build_record_dict(row, current_seq))
                elif line_type != 'MDT':
                    last_valid_seq = current_seq

    if is_in_anomaly_block and current_anomaly_group:
        first_anomaly = current_anomaly_group[0]['_seq_int']
        last_anomaly = current_anomaly_group[-1]['_seq_int']
        skipped_range = f"{first_anomaly} – {last_anomaly}" if first_anomaly != last_anomaly else str(first_anomaly)
        gaps.append({
            'from_dsn': anomaly_start_seq,
            'to_dsn': 'End of File',
            'skipped_count': len(current_anomaly_group),
            'skipped_range': skipped_range
        })

    # =========================================================
    # 3. GENERATE STRATIFICATIONS
    # =========================================================
    hold_strat = {}
    hold_type_strat = {'Bus': 0, 'Truck': 0}
    hold_wc_strat = {'HWC': 0, 'LWC': 0} if include_work_content else {}
    hold_region_strat = {'Domestic': 0, 'Export': 0}
    
    for r in hold_records:
        var_str = r.get('variant', '').strip().upper()
        if var_str.startswith(('V83', 'F83', 'M83', 'L83')):
            hold_type_strat['Bus'] += 1
            r['vehicle_type'] = 'Bus'
        else:
            hold_type_strat['Truck'] += 1
            r['vehicle_type'] = 'Truck'
            
        if include_work_content:
            wc = r.get('work_content', 'LWC')
            hold_wc_strat[wc] = hold_wc_strat.get(wc, 0) + 1
        
        reg = r.get('region', 'Export')
        hold_region_strat[reg] = hold_region_strat.get(reg, 0) + 1
        model = get_pie_model_label(r)
        hold_strat[model] = hold_strat.get(model, 0) + 1
        
    skip_strat = {}
    skip_type_strat = {'Bus': 0, 'Truck': 0}
    skip_wc_strat = {'HWC': 0, 'LWC': 0} if include_work_content else {}
    skip_region_strat = {'Domestic': 0, 'Export': 0}
    
    for r in skip_records:
        var_str = r.get('variant', '').strip().upper()
        if var_str.startswith(('V83', 'F83', 'M83', 'L83')):
            skip_type_strat['Bus'] += 1
            r['vehicle_type'] = 'Bus'
        else:
            skip_type_strat['Truck'] += 1
            r['vehicle_type'] = 'Truck'
            
        if include_work_content:
            wc = r.get('work_content', 'LWC')
            skip_wc_strat[wc] = skip_wc_strat.get(wc, 0) + 1
        
        reg = r.get('region', 'Export')
        skip_region_strat[reg] = skip_region_strat.get(reg, 0) + 1
        model = get_pie_model_label(r)
        skip_strat[model] = skip_strat.get(model, 0) + 1

    # =========================================================
    # 4. PREPARE FINAL JSON RESPONSE
    # =========================================================
    total_rows = len(original_df)
    total_hold = len(hold_records)
    total_skipped = len(skip_records)
    
    valid_seqs = df['_seq_int']
    dsn_min = int(valid_seqs.min()) if not valid_seqs.empty else 0
    dsn_max = int(valid_seqs.max()) if not valid_seqs.empty else 0
    
    return {
        'summary': {
            'dsn_min': dsn_min,
            'dsn_max': dsn_max,
            'total_in_file': total_rows,
            'total_hold': total_hold,
            'total_skipped': total_skipped,
            'hold_stratification': hold_strat,
            'skip_stratification': skip_strat,
            'hold_type_stratification': hold_type_strat,
            'skip_type_stratification': skip_type_strat,
            'hold_wc_stratification': hold_wc_strat,
            'skip_wc_stratification': skip_wc_strat,
            'hold_region_stratification': hold_region_strat,
            'skip_region_stratification': skip_region_strat,
            'shortage_parts': list(shortages.keys())
        },
        'hold_orders': hold_records,
        'skip_orders': skip_records,
        'gaps': gaps,
        'preview_columns': [str(c) for c in original_df.columns.tolist()],
        'preview_data': original_df.fillna('').astype(str).replace('nan', '').to_dict(orient='records')
    }

@app.route('/api/analyze', methods=['POST'])
def analyze():
    if 'file' not in request.files:
        return jsonify({'error': 'No main sequence file uploaded'}), 400
    
    f = request.files['file']
    if not f.filename.endswith(('.xlsx', '.xls', '.csv')):
        return jsonify({'error': 'Only .xlsx, .xls, and .csv files are supported'}), 400
        
    # Process Shortage Files & Inputs
    shortage_parts = request.form.getlist('shortage_parts')
    shortage_refs = request.form.getlist('shortage_refs')
    shortage_qtys = request.form.getlist('shortage_qtys')
    shortage_usages = request.form.getlist('shortage_usages')
    shortage_files = request.files.getlist('shortage_files')
    engine_status_file = request.files.get('engine_status_file')
    axle_status_file = request.files.get('axle_status_file')
    frame_status_file = request.files.get('frame_status_file')
    opening_hold_keys = request.form.getlist('opening_hold_keys')
    line_type = request.form.get('line_type', 'HDT')
    
    shortages = {}
    for i in range(min(len(shortage_parts), len(shortage_files))):
        part_num = shortage_parts[i].strip()
        ref_order = shortage_refs[i].strip() if i < len(shortage_refs) else ''
        try:
            qty = int(shortage_qtys[i].strip())
        except:
            qty = 0
        try:
            usage = int(shortage_usages[i].strip()) if i < len(shortage_usages) else 1
        except:
            usage = 1
        usage = max(usage, 1)
            
        file_obj = shortage_files[i]
        
        if part_num and file_obj.filename:
            try:
                df_part = parse_excel(file_obj.read())
                # Ensure Column F (Index 5) exists for Variant extraction
                if len(df_part.columns) > 5:
                    variants = df_part.iloc[:, 5].astype(str).str.strip().str.upper().tolist()
                    shortages[part_num] = {
                        'variants': set(variants),
                        'ref_order': ref_order,
                        'qty': qty,
                        'usage': usage
                    }
            except Exception as e:
                print(f"Error parsing shortage file for {part_num}: {e}")

    engine_transmission_statuses = None
    if engine_status_file and engine_status_file.filename:
        engine_transmission_statuses = {}
        try:
            df_status = parse_excel(engine_status_file.read())
            if len(df_status.columns) > 11:
                order_series = df_status.iloc[:, 2]
                engine_series = df_status.iloc[:, 9]
                transmission_series = df_status.iloc[:, 11]

                for order_value, engine_value, transmission_value in zip(order_series, engine_series, transmission_series):
                    order_key = normalize_order_key(order_value)
                    if order_key:
                        engine_transmission_statuses[order_key] = {
                            'engine_status': '' if pd.isna(engine_value) else str(engine_value).strip(),
                            'transmission_status': '' if pd.isna(transmission_value) else str(transmission_value).strip(),
                        }
        except Exception as e:
            print(f"Error parsing engine/transmission status file: {e}")

    axle_statuses = None
    if axle_status_file and axle_status_file.filename:
        try:
            axle_statuses = parse_axle_status_report(axle_status_file.read())
        except Exception as e:
            print(f"Error parsing axle status file: {e}")

    frame_statuses = None
    if frame_status_file and frame_status_file.filename:
        try:
            frame_statuses = parse_frame_status_report(frame_status_file.read())
        except Exception as e:
            print(f"Error parsing frame status file: {e}")

    try:
        df = parse_excel(f.read())
        result = analyze_sequence(df, shortages, engine_transmission_statuses, axle_statuses, frame_statuses, opening_hold_keys, line_type)
        return jsonify(result)
    except Exception as e:
        print(f"Server Error during analysis: {str(e)}")
        return jsonify({'error': f"Processing error: {str(e)}"}), 500


@app.route('/api/critical-part-details', methods=['POST'])
def critical_part_details():
    report_file = request.files.get('seven_days_report_file')
    part_number = request.form.get('part_number', '')
    part_numbers = parse_json_field('part_numbers', None)

    requested_parts = part_numbers if isinstance(part_numbers, list) else [part_number]
    requested_parts = [normalize_part_number_value(part) for part in requested_parts if normalize_part_number_value(part)]
    if not requested_parts:
        return jsonify({'error': 'Enter a part number.'}), 400
    if not report_file or not report_file.filename:
        return jsonify({'error': 'Upload the 7 days report before looking up part details.'}), 400

    try:
        details_by_part = parse_critical_parts_details(report_file.read(), requested_parts)
        if isinstance(part_numbers, list):
            return jsonify({
                'parts': [details_by_part.get(normalize_part_number_value(part).upper()) for part in requested_parts],
                'missing': [part for part in requested_parts if normalize_part_number_value(part).upper() not in details_by_part],
            })
        details = details_by_part.get(normalize_part_number_value(requested_parts[0]).upper())
        if not details:
            return jsonify({'error': 'Part number was not found in the 7 days report.'}), 404
        return jsonify(details)
    except Exception as e:
        print(f"Critical part lookup failed: {str(e)}")
        return jsonify({'error': f"Critical part lookup failed: {str(e)}"}), 500


@app.route('/api/l4-directory', methods=['POST'])
def l4_directory():
    report_file = request.files.get('seven_days_report_file')
    if not report_file or not report_file.filename:
        return jsonify({'error': 'Upload the 7 days report before loading L4 details.'}), 400

    try:
        return jsonify({'l4Directory': parse_l4_directory(report_file.read())})
    except Exception as e:
        print(f"L4 directory load failed: {str(e)}")
        return jsonify({'error': f"L4 directory load failed: {str(e)}"}), 500


@app.route('/api/constraint-backups', methods=['GET'])
def constraint_backups():
    backups = [describe_backup_folder(folder) for folder in get_backup_folders()]
    return jsonify({'backups': backups})


@app.route('/api/constraint-backups/<backup_id>', methods=['GET'])
def constraint_backup_detail(backup_id):
    safe_backup_id = sanitize_backup_name(backup_id, '')
    backup_folder = (BACKUP_DIR / safe_backup_id).resolve()

    if not safe_backup_id or BACKUP_DIR.resolve() not in backup_folder.parents:
        return jsonify({'error': 'Invalid backup folder.'}), 400
    if not backup_folder.exists() or not backup_folder.is_dir():
        return jsonify({'error': 'Backup folder was not found.'}), 404

    mapped_csv = backup_folder / 'Mapped_Day_Opening_Report.csv'
    mapped_excel = backup_folder / 'Mapped_Day_Opening_Report.xlsx'
    if mapped_csv.exists():
        mapped_df = pd.read_csv(mapped_csv)
    elif mapped_excel.exists():
        mapped_df = pd.read_excel(mapped_excel)
    else:
        return jsonify({'error': 'Mapped opening report is missing from this backup.'}), 404

    mapped_df = mapped_df.fillna('')
    mapped_columns = [str(column) for column in mapped_df.columns]
    mapped_rows = mapped_df.to_dict(orient='records')

    return jsonify({
        'backup': describe_backup_folder(backup_folder),
        'mapped_columns': mapped_columns,
        'mapped_rows': mapped_rows,
        'summary': read_json_file(backup_folder / 'Opening_Summary.json', {}),
        'status_summary': read_json_file(backup_folder / 'Status_Summary.json', {}),
        'inference_cards': read_json_file(backup_folder / 'Inference_Cards.json', []),
        'workspace_state': read_json_file(backup_folder / 'Planner_Session.json', {}),
    })


@app.route('/api/save-constraints', methods=['POST'])
def save_constraints():
    mapped_columns = parse_json_field('mapped_columns', [])
    summary = parse_json_field('summary', {})
    status_summary = parse_json_field('status_summary', {})
    inference_cards = parse_json_field('inference_cards', [])
    chart_images = parse_json_field('chart_images', [])
    workspace_state = parse_json_field('workspace_state', {})
    line_type = request.form.get('line_type', 'HDT')
    mapped_report_file = request.files.get('mapped_report_file')

    if not mapped_columns or not mapped_report_file or not mapped_report_file.filename:
        return jsonify({'error': 'No mapped report data available to save.'}), 400

    try:
        backup_date = datetime.now().strftime('%Y-%m-%d')
        backup_folder = BACKUP_DIR / backup_date
        source_folder = backup_folder / 'source_files'
        backup_folder.mkdir(parents=True, exist_ok=True)
        source_folder.mkdir(parents=True, exist_ok=True)

        mapped_df = pd.read_csv(mapped_report_file)
        mapped_columns = [str(column) for column in mapped_columns]
        for column in mapped_columns:
            if column not in mapped_df.columns:
                mapped_df[column] = ''
        mapped_df = mapped_df[mapped_columns]

        mapped_excel_path = backup_folder / 'Mapped_Day_Opening_Report.xlsx'
        mapped_csv_path = backup_folder / 'Mapped_Day_Opening_Report.csv'
        mapped_df.to_excel(mapped_excel_path, index=False)
        mapped_df.to_csv(mapped_csv_path, index=False, encoding='utf-8-sig')

        saved_files = [
            mapped_excel_path.name,
            mapped_csv_path.name,
        ]

        metadata_files = [
            ('Opening_Summary.json', summary),
            ('Status_Summary.json', status_summary),
            ('Inference_Cards.json', inference_cards),
            ('Planner_Session.json', {
                **workspace_state,
                'lineType': line_type,
                'savedAt': datetime.now().isoformat(),
            }),
        ]
        for file_name, payload in metadata_files:
            metadata_path = backup_folder / file_name
            metadata_path.write_text(json.dumps(payload, indent=2), encoding='utf-8')
            saved_files.append(metadata_path.name)

        file_groups = [
            ('opening_report_file', 'opening_report'),
            ('mod_report_file', 'mod_report'),
            ('seven_days_report_file', 'seven_days_report'),
            ('engine_status_file', 'engine_transmission_status'),
            ('axle_status_file', 'axle_status'),
            ('frame_status_file', 'frame_status'),
        ]
        for field_name, prefix in file_groups:
            saved_name = save_uploaded_file(request.files.get(field_name), source_folder, prefix)
            if saved_name:
                saved_files.append(f"source_files/{saved_name}")

        for index, shortage_file in enumerate(request.files.getlist('shortage_files'), start=1):
            saved_name = save_uploaded_file(shortage_file, source_folder, f"shortage_{index}")
            if saved_name:
                saved_files.append(f"source_files/{saved_name}")

        pdf_bytes = build_backup_pdf_bytes(line_type, summary, status_summary, inference_cards, len(mapped_df.index), chart_images)
        pdf_path = backup_folder / 'Day_Opening_Summary.pdf'
        pdf_path.write_bytes(pdf_bytes)
        saved_files.append(pdf_path.name)

        pptx_bytes = build_backup_pptx_bytes(line_type, summary, status_summary, inference_cards, len(mapped_df.index), chart_images)
        pptx_path = backup_folder / 'Day_Opening_Summary.pptx'
        pptx_path.write_bytes(pptx_bytes)
        saved_files.append(pptx_path.name)

        return jsonify({
            'message': 'Constraints backup saved successfully.',
            'folder': str(backup_folder),
            'files': saved_files,
        })
    except Exception as e:
        print(f"Server Error during backup save: {str(e)}")
        return jsonify({'error': f"Backup save failed: {str(e)}"}), 500


@app.route('/', defaults={'path': ''})
@app.route('/<path:path>')
def index(path):
    if path.startswith('api/'):
        return jsonify({'error': f'API route not found: /{path}'}), 404

    requested_file = FRONTEND_DIST_DIR / path
    if path and requested_file.is_file():
        return send_file(requested_file)

    built_index = FRONTEND_DIST_DIR / 'index.html'
    if built_index.exists():
        return send_file(built_index)

    return send_file(BASE_DIR / 'index.html')

if __name__ == '__main__':
    app.run(debug=True, port=5050)
